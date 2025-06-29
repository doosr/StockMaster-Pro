import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
from datetime import datetime, timedelta

class StockApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Système Professionnel de Gestion de Stock")
        self.geometry("1400x800")
        
        # Configuration du fichier Excel
        self.filename = "suivi_consommation.xlsx"
        self.create_template_if_needed()
        self.load_data()
        
        # Style professionnel amélioré
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure(".", background="#f0f2f5", font=("Segoe UI", 10))
        self.style.configure("TNotebook", background="#e9ecef")
        self.style.configure("TNotebook.Tab", padding=(15, 5), font=("Segoe UI", 10, "bold"))
        self.style.configure("TFrame", background="#ffffff")
        self.style.configure("Header.TLabel", background="#3b5998", foreground="white", 
                            font=("Segoe UI", 10, "bold"), padding=5)
        self.style.configure("Accent.TButton", background="#4267B2", foreground="white", 
                            font=("Segoe UI", 10, "bold"))
        self.style.configure("Critical.TLabel", background="#ffcccc", foreground="#cc0000")
        self.style.map("Accent.TButton", background=[("active", "#365899")])
        self.style.configure("Treeview", rowheight=25)
        self.style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        
        # Onglets
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Création des onglets
        self.create_consumption_tab()
        self.create_stock_tab()
        self.create_commandes_tab()
        self.create_alerts_tab()
        self.create_report_tab()
        self.create_auxiliary_tab()
        self.create_indicators_tab()
        
        # Barre de statut
        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, 
                                  anchor=tk.W, padding=5)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        self.status_var.set("Prêt | Système de Gestion de Stock")
        
        # Initialisation
        if self.colorants:
            self.combo_ref.current(0)
            self.update_stock_display()
            self.check_stock_alerts()
            self.update_report_table()
            self.update_commandes_display()
            self.update_indicators()
            self.load_auxiliary_data()

    def create_template_if_needed(self):
        """Crée un fichier Excel modèle s'il n'existe pas"""
        if not os.path.exists(self.filename):
            try:
                wb = openpyxl.Workbook()
                
                # Feuille 1: Groupe compta. Stock
                sheet1 = wb.active
                sheet1.title = "Groupe compta. Stock"
                sheet1.append(["Liste des articles"])
                sheet1.append(["Groupe compta. Stock"])
                for _ in range(5): sheet1.append([])
                sheet1.append(["", "LES COMMANDES", "CONSOMMATION"])
                for _ in range(2): sheet1.append([])
                sheet1.append(["", "N° total de cmd traitée", "=MAX(Consommation!D2:D143)"])
                sheet1.append(["", "=COUNTIF(commandes!G2:G162,\"traitée\")", "=MIN(Consommation!D2:D136)"])
                for _ in range(2): sheet1.append([])
                sheet1.append(["", "taux des comandes", ""])
                sheet1.append(["", "=B10/B16", ""])
                for _ in range(2): sheet1.append([])
                sheet1.append(["", "N total de commandes", ""])
                sheet1.append(["", "=COUNTA(commandes!A2:A148)", ""])
                
                # Feuille 2: Liste des articles2
                sheet2 = wb.create_sheet("Liste des articles2")
                headers = [
                    "ID COLORANTS", "NOM DE COLORANT", "Stock", "STOCK MIN", 
                    "CONSOMMATION", "STOCK REEL", "DATE D'ENTRE", "ALERTE DE STOCK"
                ]
                sheet2.append(headers)
                
                # Feuille 3: Consommation
                sheet3 = wb.create_sheet("Consommation")
                sheet3.append(["ID COLORANTS/NOM DE COLORANT", "DATE", "CONSOMMATION (jours)", "CONSOMMATION ( semaine)"])
                
                # Feuille 4: Consommation total par colorant
                sheet4 = wb.create_sheet("Consommation total par colorant")
                sheet4.append(["ID COLORANT", "CONSOMMATION TOTAL (mois)"])
                
                # Feuille 5: commandes
                sheet5 = wb.create_sheet("commandes")
                headers = [
                    "LES COMMANDES", "CODE COULEUR", "DATE D'ENTRE", "DATE SORTIE", 
                    "delai (jours)", "delai de traitement", "Statut", "observation"
                ]
                sheet5.append(headers)
                
                # Feuille 6: Feuil2
                sheet6 = wb.create_sheet("Feuil2")
                sheet6.append(["statut"])
                sheet6.append(["traité"])
                sheet6.append(["non traité"])
                
                # Feuille 7: Produits auxiliaires
                sheet7 = wb.create_sheet("Produits auxiliaires")
                headers = [
                    "ID PRODUIT", "NOM", "Stock", "STOCK MIN", 
                    "CONSOMMATION", "STOCK REEL", "DATE D'ENTRE", "ALERTE DE STOCK"
                ]
                sheet7.append(headers)
                
                wb.save(self.filename)
                messagebox.showinfo("Fichier créé", 
                                   "Un nouveau fichier Excel modèle a été créé.")
            except Exception as e:
                messagebox.showerror("Erreur", 
                                    f"Impossible de créer le fichier Excel:\n{str(e)}")

    def load_data(self):
        """Charge les données depuis le fichier Excel"""
        try:
            if not os.path.exists(self.filename):
                messagebox.showerror("Erreur", f"Fichier Excel introuvable: {self.filename}")
                return
            
            self.wb = openpyxl.load_workbook(self.filename, data_only=True)
            self.sheet_articles = self.wb['Liste des articles2']
            self.sheet_consommation = self.wb['Consommation']
            self.sheet_commandes = self.wb['commandes']
            self.sheet_stats = self.wb['Groupe compta. Stock']
            
            # Récupère la liste des colorants
            self.colorants = []
            self.colorant_names = {}
            self.stock_min = {}
            self.stocks = {}
            self.stock_initial = {}
            
            # Lire les valeurs de stock
            for row in range(2, self.sheet_articles.max_row + 1):
                ref = self.sheet_articles.cell(row=row, column=1).value
                if ref:
                    ref = str(ref)
                    self.colorants.append(ref)
                    
                    # Nom du colorant
                    name = self.sheet_articles.cell(row=row, column=2).value
                    self.colorant_names[ref] = name if name else ref
                    
                    # Stock minimum
                    min_val = self.sheet_articles.cell(row=row, column=4).value
                    try:
                        min_val = float(min_val) if min_val is not None else 0.0
                    except (TypeError, ValueError):
                        min_val = 0.0
                    self.stock_min[ref] = min_val
                    
                    # Stock initial
                    init_val = self.sheet_articles.cell(row=row, column=3).value
                    try:
                        init_val = float(init_val) if init_val is not None else 0.0
                    except (TypeError, ValueError):
                        init_val = 0.0
                    self.stock_initial[ref] = init_val
                    
                    # Stock réel
                    stock_val = self.sheet_articles.cell(row=row, column=6).value
                    try:
                        stock_val = float(stock_val) if stock_val is not None else 0.0
                    except (TypeError, ValueError):
                        stock_val = 0.0
                    self.stocks[ref] = stock_val
            
            # Charger l'historique des consommations
            self.consumption_history = []
            for row in range(2, self.sheet_consommation.max_row + 1):
                ref = self.sheet_consommation.cell(row=row, column=1).value
                date_val = self.sheet_consommation.cell(row=row, column=2).value
                qty = self.sheet_consommation.cell(row=row, column=3).value
                
                if ref and date_val and qty:
                    try:
                        if isinstance(date_val, datetime):
                            date_str = date_val.strftime('%Y-%m-%d')
                        else:
                            date_str = str(date_val)
                        
                        self.consumption_history.append({
                            'ref': str(ref),
                            'date': date_str,
                            'qty': float(qty),
                            'id': row  # Stocker l'ID de ligne pour les modifications
                        })
                    except:
                        continue
            
            # Charger les commandes
            self.commandes = []
            for row in range(2, self.sheet_commandes.max_row + 1):
                cmd_ref = self.sheet_commandes.cell(row=row, column=1).value
                if cmd_ref:
                    commande = {
                        'ref': cmd_ref,
                        'code': self.sheet_commandes.cell(row=row, column=2).value,
                        'date_entree': self.sheet_commandes.cell(row=row, column=3).value,
                        'date_sortie': self.sheet_commandes.cell(row=row, column=4).value,
                        'delai': self.sheet_commandes.cell(row=row, column=5).value,
                        'statut': self.sheet_commandes.cell(row=row, column=7).value,
                        'observation': self.sheet_commandes.cell(row=row, column=8).value,
                        'id': row  # Stocker l'ID de ligne pour les modifications
                    }
                    self.commandes.append(commande)
            
            # Charger les statistiques de commandes
            self.total_commandes = self.get_cell_value('Groupe compta. Stock', 'B16')  # N total de commandes
            self.commandes_traitees = self.get_cell_value('Groupe compta. Stock', 'B10')  # N° total de cmd traitée
            self.taux_commandes = self.get_cell_value('Groupe compta. Stock', 'B13')  # taux des commandes
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de charger le fichier Excel:\n{str(e)}")
            self.colorants = []
            self.stock_min = {}
            self.stocks = {}
            self.stock_initial = {}
            self.consumption_history = []
            self.commandes = []
            self.total_commandes = 0
            self.commandes_traitees = 0
            self.taux_commandes = 0.0
    
    def get_cell_value(self, sheet_name, cell_ref):
        """Récupère la valeur d'une cellule par référence"""
        try:
            sheet = self.wb[sheet_name]
            return sheet[cell_ref].value
        except:
            return 0

    def create_consumption_tab(self):
        """Crée l'onglet de consommation"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 Consommation")
        
        # Frame principale avec padding
        main_frame = ttk.Frame(tab, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Groupe: Saisie de consommation avec style amélioré
        group = ttk.LabelFrame(main_frame, text="➕ Enregistrer une nouvelle consommation")
        group.pack(fill=tk.X, pady=(0, 15))
        
        # Formulaire
        form_frame = ttk.Frame(group, padding=10)
        form_frame.pack(fill=tk.X)
        
        # Type de produit
        ttk.Label(form_frame, text="Type de produit:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.product_type = ttk.Combobox(form_frame, values=["Colorant", "Produit auxiliaire"], state="readonly", width=15)
        self.product_type.current(0)
        self.product_type.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        self.product_type.bind("<<ComboboxSelected>>", self.update_product_list)
        
        # Référence produit
        ttk.Label(form_frame, text="Référence:", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.combo_ref = ttk.Combobox(form_frame, state="readonly", width=30)
        self.combo_ref.grid(row=1, column=1, sticky="we", padx=5, pady=5)
        self.combo_ref.bind("<<ComboboxSelected>>", lambda e: self.update_stock_display())
        
        # Consommation
        ttk.Label(form_frame, text="Consommation (kg):", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.entry_consommation = ttk.Entry(form_frame, width=15)
        self.entry_consommation.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        # Date
        ttk.Label(form_frame, text="Date:", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.entry_date = ttk.Entry(form_frame, width=15)
        self.entry_date.insert(0, datetime.today().strftime('%Y-%m-%d'))
        self.entry_date.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Stock réel
        ttk.Label(form_frame, text="Stock Réel Actuel:", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="e", padx=5, pady=5)
        self.label_stock = ttk.Label(form_frame, text="0.0 kg", font=("Segoe UI", 10, "bold"), foreground="#1a73e8")
        self.label_stock.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        
        # Bouton d'enregistrement
        btn_frame = ttk.Frame(group, padding=10)
        btn_frame.pack(fill=tk.X)
        self.btn_save = ttk.Button(btn_frame, text="💾 Enregistrer Consommation", 
                                  command=self.save_consumption, style="Accent.TButton")
        self.btn_save.pack(pady=5)
        
        # Historique des consommations
        history_group = ttk.LabelFrame(main_frame, text="🕒 Historique des Consommations")
        history_group.pack(fill=tk.BOTH, expand=True, pady=(15, 0))
        
        # Barre d'outils pour l'historique
        toolbar_frame = ttk.Frame(history_group, padding=5)
        toolbar_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(toolbar_frame, text="🔄 Actualiser", command=self.update_history_tree).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="✏️ Modifier", command=self.edit_consumption).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="🗑️ Supprimer", command=self.delete_consumption).pack(side=tk.LEFT, padx=2)
        
        # Tableau d'historique
        columns = ("date", "ref", "name", "qty", "type", "id")
        self.history_tree = ttk.Treeview(history_group, columns=columns, show="headings", selectmode="browse")
        
        # Configuration des colonnes
        self.history_tree.heading("date", text="Date")
        self.history_tree.heading("ref", text="Référence")
        self.history_tree.heading("name", text="Nom")
        self.history_tree.heading("qty", text="Quantité (kg)")
        self.history_tree.heading("type", text="Type")
        self.history_tree.heading("id", text="ID")
        
        self.history_tree.column("date", width=120, anchor="center")
        self.history_tree.column("ref", width=120, anchor="center")
        self.history_tree.column("name", width=250, anchor="w")
        self.history_tree.column("qty", width=100, anchor="e")
        self.history_tree.column("type", width=120, anchor="center")
        self.history_tree.column("id", width=50, anchor="center", stretch=False)
        
        # Masquer la colonne ID
        self.history_tree["displaycolumns"] = ("date", "ref", "name", "qty", "type")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(history_group, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement des éléments
        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.update_history_tree()
        self.update_product_list()

    def update_product_list(self, event=None):
        """Met à jour la liste des produits selon le type sélectionné"""
        product_type = self.product_type.get()
        
        if product_type == "Colorant":
            self.combo_ref['values'] = self.colorants
            if self.colorants:
                self.combo_ref.current(0)
        elif product_type == "Produit auxiliaire":
            # Charger les produits auxiliaires
            try:
                sheet = self.wb["Produits auxiliaires"]
                products = []
                for row in range(2, sheet.max_row + 1):
                    ref = sheet.cell(row=row, column=1).value
                    if ref:
                        products.append(str(ref))
                self.combo_ref['values'] = products
                if products:
                    self.combo_ref.current(0)
            except:
                self.combo_ref['values'] = []
        
        self.update_stock_display()

    def create_stock_tab(self):
        """Crée l'onglet de gestion de stock"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📦 Gestion Stock")
        
        # Frame principale avec padding
        main_frame = ttk.Frame(tab, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame avec deux colonnes
        dual_frame = ttk.Frame(main_frame)
        dual_frame.pack(fill=tk.BOTH, expand=True)
        
        # Colonne gauche - Mise à jour du stock
        left_frame = ttk.LabelFrame(dual_frame, text="🔄 Mise à jour du Stock")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10), pady=5)
        
        # Formulaire
        form_frame = ttk.Frame(left_frame, padding=10)
        form_frame.pack(fill=tk.X)
        
        # Type de produit
        ttk.Label(form_frame, text="Type de produit:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.stock_product_type = ttk.Combobox(form_frame, values=["Colorant", "Produit auxiliaire"], state="readonly", width=15)
        self.stock_product_type.current(0)
        self.stock_product_type.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        self.stock_product_type.bind("<<ComboboxSelected>>", self.update_stock_product_list)
        
        # Référence produit
        ttk.Label(form_frame, text="Référence:", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.combo_stock_ref = ttk.Combobox(form_frame, state="readonly", width=30)
        self.combo_stock_ref.grid(row=1, column=1, sticky="we", padx=5, pady=5)
        self.combo_stock_ref.bind("<<ComboboxSelected>>", lambda e: self.update_stock_info())
        
        # Stock actuel
        ttk.Label(form_frame, text="Stock Initial Actuel:", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.label_current_stock = ttk.Label(form_frame, text="0.0 kg", font=("Segoe UI", 10, "bold"), foreground="#1a73e8")
        self.label_current_stock.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        # Nouveau stock
        ttk.Label(form_frame, text="Nouveau Stock Initial:", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.entry_new_stock = ttk.Entry(form_frame, width=15)
        self.entry_new_stock.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Bouton de mise à jour
        btn_frame = ttk.Frame(left_frame, padding=10)
        btn_frame.pack(fill=tk.X)
        self.btn_update_stock = ttk.Button(btn_frame, text="🔄 Mettre à Jour le Stock", 
                                         command=self.update_initial_stock, style="Accent.TButton")
        self.btn_update_stock.pack(pady=5)
        
        # Colonne droite - Ajout de nouveau produit
        right_frame = ttk.LabelFrame(dual_frame, text="➕ Ajouter un Nouveau Produit")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=5)
        
        # Formulaire
        new_form = ttk.Frame(right_frame, padding=10)
        new_form.pack(fill=tk.X)
        
        # Type de produit
        ttk.Label(new_form, text="Type de produit:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.new_product_type = ttk.Combobox(new_form, values=["Colorant", "Produit auxiliaire"], state="readonly", width=15)
        self.new_product_type.current(0)
        self.new_product_type.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        # Champs pour nouveau produit
        ttk.Label(new_form, text="Référence:", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.entry_new_ref = ttk.Entry(new_form, width=30)
        self.entry_new_ref.grid(row=1, column=1, sticky="we", padx=5, pady=5)
        
        ttk.Label(new_form, text="Nom:", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.entry_new_name = ttk.Entry(new_form, width=30)
        self.entry_new_name.grid(row=2, column=1, sticky="we", padx=5, pady=5)
        
        ttk.Label(new_form, text="Stock Initial:", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.entry_new_init_stock = ttk.Entry(new_form, width=15)
        self.entry_new_init_stock.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        self.entry_new_init_stock.insert(0, "0.0")
        
        ttk.Label(new_form, text="Stock Minimal:", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="e", padx=5, pady=5)
        self.entry_new_min_stock = ttk.Entry(new_form, width=15)
        self.entry_new_min_stock.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        self.entry_new_min_stock.insert(0, "0.0")
        
        # Bouton d'ajout
        btn_frame = ttk.Frame(right_frame, padding=10)
        btn_frame.pack(fill=tk.X)
        self.btn_add_colorant = ttk.Button(btn_frame, text="➕ Ajouter Nouveau Produit", 
                                         command=self.add_new_product, style="Accent.TButton")
        self.btn_add_colorant.pack(pady=5)
        
        self.update_stock_product_list()

    def update_stock_product_list(self, event=None):
        """Met à jour la liste des produits pour la gestion de stock"""
        product_type = self.stock_product_type.get()
        
        if product_type == "Colorant":
            self.combo_stock_ref['values'] = self.colorants
            if self.colorants:
                self.combo_stock_ref.current(0)
        elif product_type == "Produit auxiliaire":
            # Charger les produits auxiliaires
            try:
                sheet = self.wb["Produits auxiliaires"]
                products = []
                for row in range(2, sheet.max_row + 1):
                    ref = sheet.cell(row=row, column=1).value
                    if ref:
                        products.append(str(ref))
                self.combo_stock_ref['values'] = products
                if products:
                    self.combo_stock_ref.current(0)
            except:
                self.combo_stock_ref['values'] = []
        
        self.update_stock_info()

    def create_commandes_tab(self):
        """Crée l'onglet de gestion des commandes"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Commandes")
        
        # Frame principale
        main_frame = ttk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Statistiques des commandes
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(stats_frame, text="Total Commandes:").pack(side=tk.LEFT, padx=5)
        self.total_cmd_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.total_cmd_var, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(stats_frame, text="Commandes Traitées:").pack(side=tk.LEFT, padx=5)
        self.traitees_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.traitees_var, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(stats_frame, text="Taux Traitement:").pack(side=tk.LEFT, padx=5)
        self.taux_var = tk.StringVar(value="0.0%")
        ttk.Label(stats_frame, textvariable=self.taux_var, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        # Boutons de gestion des commandes
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(btn_frame, text="Nouvelle Commande", command=self.ajouter_commande, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Modifier", command=self.modifier_commande).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Marquer comme Traitée", command=self.marquer_traitee).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Supprimer", command=self.supprimer_commande).pack(side=tk.LEFT, padx=5)
        
        # Liste des commandes
        commandes_frame = ttk.LabelFrame(main_frame, text="Liste des Commandes")
        commandes_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ("ref", "code", "date_entree", "date_sortie", "delai", "statut")
        self.commandes_tree = ttk.Treeview(commandes_frame, columns=columns, show="headings")
        
        # Configuration des colonnes
        self.commandes_tree.heading("ref", text="Référence")
        self.commandes_tree.heading("code", text="Code Couleur")
        self.commandes_tree.heading("date_entree", text="Date Entrée")
        self.commandes_tree.heading("date_sortie", text="Date Sortie")
        self.commandes_tree.heading("delai", text="Délai (jours)")
        self.commandes_tree.heading("statut", text="Statut")
        
        self.commandes_tree.column("ref", width=100, anchor="center")
        self.commandes_tree.column("code", width=150, anchor="w")
        self.commandes_tree.column("date_entree", width=120, anchor="center")
        self.commandes_tree.column("date_sortie", width=120, anchor="center")
        self.commandes_tree.column("delai", width=100, anchor="center")
        self.commandes_tree.column("statut", width=100, anchor="center")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(commandes_frame, orient="vertical", command=self.commandes_tree.yview)
        self.commandes_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement des éléments
        self.commandes_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Mettre à jour l'affichage
        self.update_commandes_display()

    def create_alerts_tab(self):
        """Crée l'onglet d'alertes"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Alertes Stock")
        
        # Frame principale
        main_frame = ttk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Groupe: Alertes
        alert_group = ttk.LabelFrame(main_frame, text="Alertes de Stock - Niveau Critique")
        alert_group.pack(fill=tk.BOTH, expand=True)
        
        # Liste d'alertes
        self.alert_list = tk.Listbox(alert_group, font=("Arial", 10), bg="#ffffff", selectbackground="#e0e0e0")
        self.alert_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Bouton de rafraîchissement
        btn_frame = ttk.Frame(alert_group)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_refresh = ttk.Button(btn_frame, text="Actualiser les Alertes", command=self.check_stock_alerts)
        btn_refresh.pack(pady=5)

    def create_report_tab(self):
        """Crée l'onglet de rapports"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Rapports")
        
        # Frame principale
        main_frame = ttk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Groupe: Rapport de stock
        report_group = ttk.LabelFrame(main_frame, text="Rapport Complet des Stocks")
        report_group.pack(fill=tk.BOTH, expand=True)
        
        # Tableau de rapport
        columns = ("ref", "name", "stock_init", "stock_reel", "stock_min", "status", "type")
        self.report_tree = ttk.Treeview(report_group, columns=columns, show="headings")
        
        # Configuration des colonnes
        self.report_tree.heading("ref", text="Référence")
        self.report_tree.heading("name", text="Nom")
        self.report_tree.heading("stock_init", text="Stock Initial")
        self.report_tree.heading("stock_reel", text="Stock Réel")
        self.report_tree.heading("stock_min", text="Stock Minimal")
        self.report_tree.heading("status", text="Statut")
        self.report_tree.heading("type", text="Type")
        
        self.report_tree.column("ref", width=120, anchor="center")
        self.report_tree.column("name", width=200, anchor="w")
        self.report_tree.column("stock_init", width=100, anchor="e")
        self.report_tree.column("stock_reel", width=100, anchor="e")
        self.report_tree.column("stock_min", width=100, anchor="e")
        self.report_tree.column("status", width=100, anchor="center")
        self.report_tree.column("type", width=120, anchor="center")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(report_group, orient="vertical", command=self.report_tree.yview)
        self.report_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement des éléments
        self.report_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bouton d'export
        btn_frame = ttk.Frame(report_group)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_export = ttk.Button(btn_frame, text="Exporter vers Excel", command=self.export_to_excel)
        btn_export.pack(pady=5)

    def update_stock_display(self):
        """Met à jour l'affichage du stock réel"""
        product_type = self.product_type.get()
        ref = self.combo_ref.get()
        
        if product_type == "Colorant" and ref:
            self.label_stock.config(text=f"{self.stocks.get(ref, 0.0):.2f} kg")
        elif product_type == "Produit auxiliaire" and ref:
            # Récupérer le stock du produit auxiliaire
            try:
                sheet = self.wb["Produits auxiliaires"]
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == ref:
                        stock_val = sheet.cell(row=row, column=6).value
                        try:
                            stock_val = float(stock_val) if stock_val is not None else 0.0
                        except:
                            stock_val = 0.0
                        self.label_stock.config(text=f"{stock_val:.2f} kg")
                        break
            except:
                self.label_stock.config(text="0.0 kg")

    def update_stock_info(self):
        """Met à jour l'affichage du stock dans l'onglet gestion de stock"""
        product_type = self.stock_product_type.get()
        ref = self.combo_stock_ref.get()
        
        if product_type == "Colorant" and ref:
            self.label_current_stock.config(text=f"{self.stock_initial.get(ref, 0.0):.2f} kg")
            self.entry_new_stock.delete(0, tk.END)
            self.entry_new_stock.insert(0, str(self.stock_initial.get(ref, 0.0)))
        elif product_type == "Produit auxiliaire" and ref:
            # Récupérer le stock initial du produit auxiliaire
            try:
                sheet = self.wb["Produits auxiliaires"]
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == ref:
                        stock_val = sheet.cell(row=row, column=3).value
                        try:
                            stock_val = float(stock_val) if stock_val is not None else 0.0
                        except:
                            stock_val = 0.0
                        self.label_current_stock.config(text=f"{stock_val:.2f} kg")
                        self.entry_new_stock.delete(0, tk.END)
                        self.entry_new_stock.insert(0, str(stock_val))
                        break
            except:
                self.label_current_stock.config(text="0.0 kg")

    def update_history_tree(self):
        """Met à jour l'arbre d'historique des consommations"""
        # Effacer les anciennes entrées
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        # Trier par date récente
        recent_history = sorted(self.consumption_history, 
                               key=lambda x: x['date'], 
                               reverse=True)[:20]
        
        # Ajouter les nouvelles entrées
        for item in recent_history:
            ref = item['ref']
            # Déterminer le type de produit
            product_type = "Colorant" if ref in self.colorants else "Produit auxiliaire"
            
            # Trouver le nom du produit
            if product_type == "Colorant":
                name = self.colorant_names.get(ref, ref)
            else:
                name = ref  # Pour les produits auxiliaires, on utilise la référence comme nom par défaut
                try:
                    sheet = self.wb["Produits auxiliaires"]
                    for row in range(2, sheet.max_row + 1):
                        if str(sheet.cell(row=row, column=1).value) == ref:
                            name = sheet.cell(row=row, column=2).value or ref
                            break
                except:
                    pass
            
            self.history_tree.insert("", "end", values=(
                item['date'], 
                ref, 
                name, 
                f"{item['qty']:.2f}",
                product_type,
                item['id']
            ))

    def update_report_table(self):
        """Met à jour le tableau de rapport de stock"""
        # Effacer les anciennes entrées
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Ajouter les colorants
        for ref in self.colorants:
            name = self.colorant_names.get(ref, ref)
            stock_init = self.stock_initial.get(ref, 0.0)
            stock_reel = self.stocks.get(ref, 0.0)
            stock_min = self.stock_min.get(ref, 0.0)
            
            status = "CRITIQUE" if stock_reel < stock_min else "OK"
            
            self.report_tree.insert("", "end", values=(
                ref, 
                name, 
                f"{stock_init:.2f}", 
                f"{stock_reel:.2f}", 
                f"{stock_min:.2f}", 
                status,
                "Colorant"
            ))
            
            # Colorer les lignes critiques
            if stock_reel < stock_min:
                self.report_tree.item(self.report_tree.get_children()[-1], tags=("critical",))
        
        # Ajouter les produits auxiliaires
        try:
            sheet = self.wb["Produits auxiliaires"]
            for row in range(2, sheet.max_row + 1):
                ref = sheet.cell(row=row, column=1).value
                if ref:
                    ref = str(ref)
                    name = sheet.cell(row=row, column=2).value or ref
                    
                    try:
                        stock_init = float(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3).value is not None else 0.0
                    except:
                        stock_init = 0.0
                    
                    try:
                        stock_reel = float(sheet.cell(row=row, column=6).value) if sheet.cell(row=row, column=6).value is not None else 0.0
                    except:
                        stock_reel = 0.0
                    
                    try:
                        stock_min = float(sheet.cell(row=row, column=4).value) if sheet.cell(row=row, column=4).value is not None else 0.0
                    except:
                        stock_min = 0.0
                    
                    status = "CRITIQUE" if stock_reel < stock_min else "OK"
                    
                    self.report_tree.insert("", "end", values=(
                        ref, 
                        name, 
                        f"{stock_init:.2f}", 
                        f"{stock_reel:.2f}", 
                        f"{stock_min:.2f}", 
                        status,
                        "Produit auxiliaire"
                    ))
                    
                    # Colorer les lignes critiques
                    if stock_reel < stock_min:
                        self.report_tree.item(self.report_tree.get_children()[-1], tags=("critical",))
        except:
            pass
        
        # Configurer le style pour les lignes critiques
        self.report_tree.tag_configure("critical", background="#ffcccc")

    def check_stock_alerts(self):
        """Vérifie les alertes de stock et les affiche en rouge"""
        self.alert_list.delete(0, tk.END)
        alerts = []
        
        # Vérifier les colorants
        for ref, min_val in self.stock_min.items():
            current_stock = self.stocks.get(ref, 0.0)
            
            # Vérification de l'alerte
            if current_stock < min_val:
                name = self.colorant_names.get(ref, ref)
                alerts.append(f"Colorant: {ref} - {name}: Stock actuel {current_stock:.2f} kg (Min: {min_val:.2f} kg)")
        
        # Vérifier les produits auxiliaires
        try:
            sheet = self.wb["Produits auxiliaires"]
            for row in range(2, sheet.max_row + 1):
                ref = sheet.cell(row=row, column=1).value
                if ref:
                    ref = str(ref)
                    name = sheet.cell(row=row, column=2).value or ref
                    
                    try:
                        min_val = float(sheet.cell(row=row, column=4).value) if sheet.cell(row=row, column=4).value is not None else 0.0
                    except:
                        min_val = 0.0
                    
                    try:
                        current_stock = float(sheet.cell(row=row, column=6).value) if sheet.cell(row=row, column=6).value is not None else 0.0
                    except:
                        current_stock = 0.0
                    
                    if current_stock < min_val:
                        alerts.append(f"Produit auxiliaire: {ref} - {name}: Stock actuel {current_stock:.2f} kg (Min: {min_val:.2f} kg)")
        except:
            pass
        
        if alerts:
            for alert in alerts:
                self.alert_list.insert(tk.END, alert)
                self.alert_list.itemconfig(tk.END, fg="red")
        else:
            self.alert_list.insert(tk.END, "Aucune alerte de stock - tous les niveaux sont suffisants")
            self.alert_list.itemconfig(tk.END, fg="green")

    def update_commandes_display(self):
        """Met à jour l'affichage des commandes et des statistiques"""
        # Effacer les anciennes entrées
        for item in self.commandes_tree.get_children():
            self.commandes_tree.delete(item)
        
        # Ajouter les commandes
        for cmd in self.commandes:
            date_entree = cmd['date_entree']
            date_sortie = cmd['date_sortie']
            
            # Formater les dates
            if isinstance(date_entree, datetime):
                date_entree = date_entree.strftime('%Y-%m-%d')
            
            if isinstance(date_sortie, datetime):
                date_sortie = date_sortie.strftime('%Y-%m-%d')
            
            self.commandes_tree.insert("", "end", values=(
                cmd['ref'],
                cmd['code'],
                date_entree or "",
                date_sortie or "",
                cmd['delai'] or "",
                cmd['statut'] or ""
            ))
        
        # Mettre à jour les statistiques
        self.total_commandes = len(self.commandes)
        self.commandes_traitees = sum(1 for cmd in self.commandes if cmd['statut'] and "traitée" in cmd['statut'].lower())
        
        # Calculer le taux de commandes traitées
        if self.total_commandes > 0:
            self.taux_commandes = (self.commandes_traitees / self.total_commandes) * 100
        else:
            self.taux_commandes = 0.0
        
        # Mettre à jour les variables d'affichage
        self.total_cmd_var.set(str(self.total_commandes))
        self.traitees_var.set(str(self.commandes_traitees))
        self.taux_var.set(f"{self.taux_commandes:.1f}%")

    def ajouter_commande(self):
        """Ouvre une fenêtre pour ajouter une nouvelle commande"""
        dialog = tk.Toplevel(self)
        dialog.title("Nouvelle Commande")
        dialog.geometry("400x300")
        dialog.transient(self)
        dialog.grab_set()
        
        # Formulaire
        form_frame = ttk.Frame(dialog, padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Référence commande
        ttk.Label(form_frame, text="Référence Commande:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        entry_ref = ttk.Entry(form_frame)
        entry_ref.grid(row=0, column=1, sticky="we", padx=5, pady=5)
        
        # Code couleur
        ttk.Label(form_frame, text="Code Couleur:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        entry_code = ttk.Entry(form_frame)
        entry_code.grid(row=1, column=1, sticky="we", padx=5, pady=5)
        
        # Date entrée
        ttk.Label(form_frame, text="Date Entrée:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        entry_date_entree = ttk.Entry(form_frame)
        entry_date_entree.insert(0, datetime.today().strftime('%Y-%m-%d'))
        entry_date_entree.grid(row=2, column=1, sticky="we", padx=5, pady=5)
        
        # Statut
        ttk.Label(form_frame, text="Statut:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
        combo_statut = ttk.Combobox(form_frame, values=["En Attente", "Traitée", "Annulée"], state="readonly")
        combo_statut.set("En Attente")
        combo_statut.grid(row=3, column=1, sticky="we", padx=5, pady=5)
        
        # Observation
        ttk.Label(form_frame, text="Observation:").grid(row=4, column=0, sticky="e", padx=5, pady=5)
        entry_observation = ttk.Entry(form_frame)
        entry_observation.grid(row=4, column=1, sticky="we", padx=5, pady=5)
        
        # Boutons
        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(fill=tk.X)
        
        def save_cmd():
            ref = entry_ref.get().strip()
            code = entry_code.get().strip()
            date_entree = entry_date_entree.get().strip()
            statut = combo_statut.get()
            observation = entry_observation.get().strip()
            
            if not ref:
                messagebox.showwarning("Erreur", "La référence de commande est obligatoire")
                return
                
            # Vérification de la date
            try:
                datetime.strptime(date_entree, '%Y-%m-%d')
            except ValueError:
                messagebox.showwarning("Erreur", "Format de date invalide. Utilisez AAAA-MM-JJ")
                return
                
            # Créer la commande
            commande = {
                'ref': ref,
                'code': code,
                'date_entree': date_entree,
                'date_sortie': "",
                'delai': "",
                'statut': statut,
                'observation': observation
            }
            
            # Si la commande est marquée comme traitée, ajouter la date de sortie
            if "traitée" in statut.lower():
                commande['date_sortie'] = datetime.today().strftime('%Y-%m-%d')
                try:
                    date_entree_dt = datetime.strptime(date_entree, '%Y-%m-%d')
                    date_sortie_dt = datetime.today()
                    delai = (date_sortie_dt - date_entree_dt).days
                    commande['delai'] = delai
                except:
                    commande['delai'] = 0
            
            # Ajouter à la liste
            self.commandes.append(commande)
            
            # Mettre à jour l'affichage
            self.update_commandes_display()
            
            # Sauvegarder dans Excel
            self.save_commandes_to_excel()
            
            dialog.destroy()
        
        ttk.Button(btn_frame, text="Enregistrer", command=save_cmd, style="Accent.TButton").pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Annuler", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def modifier_commande(self):
        """Modifie la commande sélectionnée"""
        selected = self.commandes_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une commande")
            return
            
        item = self.commandes_tree.item(selected[0])
        values = item['values']
        ref = values[0]
        
        # Trouver la commande
        commande = next((cmd for cmd in self.commandes if cmd['ref'] == ref), None)
        if not commande:
            messagebox.showwarning("Erreur", "Commande introuvable")
            return
            
        dialog = tk.Toplevel(self)
        dialog.title("Modifier Commande")
        dialog.geometry("400x300")
        dialog.transient(self)
        dialog.grab_set()
        
        # Formulaire
        form_frame = ttk.Frame(dialog, padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Référence commande
        ttk.Label(form_frame, text="Référence Commande:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        entry_ref = ttk.Entry(form_frame)
        entry_ref.insert(0, commande['ref'])
        entry_ref.config(state='readonly')
        entry_ref.grid(row=0, column=1, sticky="we", padx=5, pady=5)
        
        # Code couleur
        ttk.Label(form_frame, text="Code Couleur:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        entry_code = ttk.Entry(form_frame)
        entry_code.insert(0, commande['code'] or "")
        entry_code.grid(row=1, column=1, sticky="we", padx=5, pady=5)
        
        # Date entrée
        ttk.Label(form_frame, text="Date Entrée:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        entry_date_entree = ttk.Entry(form_frame)
        if isinstance(commande['date_entree'], datetime):
            entry_date_entree.insert(0, commande['date_entree'].strftime('%Y-%m-%d'))
        else:
            entry_date_entree.insert(0, commande['date_entree'] or "")
        entry_date_entree.grid(row=2, column=1, sticky="we", padx=5, pady=5)
        
        # Date sortie
        ttk.Label(form_frame, text="Date Sortie:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
        entry_date_sortie = ttk.Entry(form_frame)
        if isinstance(commande['date_sortie'], datetime):
            entry_date_sortie.insert(0, commande['date_sortie'].strftime('%Y-%m-%d'))
        else:
            entry_date_sortie.insert(0, commande['date_sortie'] or "")
        entry_date_sortie.grid(row=3, column=1, sticky="we", padx=5, pady=5)
        
        # Statut
        ttk.Label(form_frame, text="Statut:").grid(row=4, column=0, sticky="e", padx=5, pady=5)
        combo_statut = ttk.Combobox(form_frame, values=["En Attente", "Traitée", "Annulée"], state="readonly")
        combo_statut.set(commande['statut'] or "En Attente")
        combo_statut.grid(row=4, column=1, sticky="we", padx=5, pady=5)
        
        # Observation
        ttk.Label(form_frame, text="Observation:").grid(row=5, column=0, sticky="e", padx=5, pady=5)
        entry_observation = ttk.Entry(form_frame)
        entry_observation.insert(0, commande['observation'] or "")
        entry_observation.grid(row=5, column=1, sticky="we", padx=5, pady=5)
        
        # Boutons
        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(fill=tk.X)
        
        def save_cmd():
            code = entry_code.get().strip()
            date_entree = entry_date_entree.get().strip()
            date_sortie = entry_date_sortie.get().strip()
            statut = combo_statut.get()
            observation = entry_observation.get().strip()
            
            # Vérification des dates
            try:
                if date_entree:
                    datetime.strptime(date_entree, '%Y-%m-%d')
                if date_sortie:
                    datetime.strptime(date_sortie, '%Y-%m-%d')
            except ValueError:
                messagebox.showwarning("Erreur", "Format de date invalide. Utilisez AAAA-MM-JJ")
                return
                
            # Mettre à jour la commande
            commande['code'] = code
            commande['date_entree'] = date_entree
            commande['date_sortie'] = date_sortie
            commande['statut'] = statut
            commande['observation'] = observation
            
            # Calculer le délai si les deux dates sont présentes
            if date_entree and date_sortie:
                try:
                    date_entree_dt = datetime.strptime(date_entree, '%Y-%m-%d')
                    date_sortie_dt = datetime.strptime(date_sortie, '%Y-%m-%d')
                    delai = (date_sortie_dt - date_entree_dt).days
                    commande['delai'] = delai
                except:
                    commande['delai'] = 0
            
            # Mettre à jour l'affichage
            self.update_commandes_display()
            
            # Sauvegarder dans Excel
            self.save_commandes_to_excel()
            
            dialog.destroy()
        
        ttk.Button(btn_frame, text="Enregistrer", command=save_cmd, style="Accent.TButton").pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Annuler", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def marquer_traitee(self):
        """Marque la commande sélectionnée comme traitée"""
        selected = self.commandes_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une commande")
            return
            
        item = self.commandes_tree.item(selected[0])
        values = item['values']
        ref = values[0]
        
        # Trouver la commande
        commande = next((cmd for cmd in self.commandes if cmd['ref'] == ref), None)
        if not commande:
            messagebox.showwarning("Erreur", "Commande introuvable")
            return
            
        # Vérifier si elle est déjà traitée
        if commande['statut'] and "traitée" in commande['statut'].lower():
            messagebox.showinfo("Info", "Cette commande est déjà marquée comme traitée")
            return
            
        # Mettre à jour la commande
        commande['statut'] = "Traitée"
        commande['date_sortie'] = datetime.today().strftime('%Y-%m-%d')
        
        # Calculer le délai
        try:
            if isinstance(commande['date_entree'], datetime):
                date_entree_dt = commande['date_entree']
            else:
                date_entree_dt = datetime.strptime(commande['date_entree'], '%Y-%m-%d')
                
            date_sortie_dt = datetime.today()
            delai = (date_sortie_dt - date_entree_dt).days
            commande['delai'] = delai
        except:
            commande['delai'] = 0
            
        # Mettre à jour l'affichage
        self.update_commandes_display()
        
        # Sauvegarder dans Excel
        self.save_commandes_to_excel()
        
        messagebox.showinfo("Succès", f"Commande {ref} marquée comme traitée")
        self.status_var.set(f"Commande {ref} marquée comme traitée")

    def supprimer_commande(self):
        """Supprime la commande sélectionnée"""
        selected = self.commandes_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une commande")
            return
            
        item = self.commandes_tree.item(selected[0])
        values = item['values']
        ref = values[0]
        
        # Confirmation
        if not messagebox.askyesno("Confirmation", f"Voulez-vous vraiment supprimer la commande {ref}?"):
            return
            
        # Supprimer la commande
        self.commandes = [cmd for cmd in self.commandes if cmd['ref'] != ref]
        
        # Mettre à jour l'affichage
        self.update_commandes_display()
        
        # Sauvegarder dans Excel
        self.save_commandes_to_excel()
        
        messagebox.showinfo("Succès", f"Commande {ref} supprimée")
        self.status_var.set(f"Commande {ref} supprimée")

    def save_commandes_to_excel(self):
        """Sauvegarde les commandes dans le fichier Excel"""
        try:
            # Effacer les anciennes données
            for row in range(2, self.sheet_commandes.max_row + 1):
                for col in range(1, 9):
                    self.sheet_commandes.cell(row=row, column=col).value = None
            
            # Écrire les nouvelles données
            for i, cmd in enumerate(self.commandes, start=2):
                self.sheet_commandes.cell(row=i, column=1, value=cmd['ref'])
                self.sheet_commandes.cell(row=i, column=2, value=cmd['code'])
                self.sheet_commandes.cell(row=i, column=3, value=cmd['date_entree'])
                self.sheet_commandes.cell(row=i, column=4, value=cmd['date_sortie'])
                self.sheet_commandes.cell(row=i, column=5, value=cmd['delai'])
                self.sheet_commandes.cell(row=i, column=7, value=cmd['statut'])
                self.sheet_commandes.cell(row=i, column=8, value=cmd['observation'])
            
            # Mettre à jour les statistiques dans la feuille "Groupe compta. Stock"
            self.sheet_stats['B10'] = self.commandes_traitees  # N° total de cmd traitée
            self.sheet_stats['B16'] = self.total_commandes    # N total de commandes
            
            if self.total_commandes > 0:
                taux = self.commandes_traitees / self.total_commandes
                self.sheet_stats['B13'] = taux  # taux des commandes
            
            # Sauvegarder le fichier
            self.wb.save(self.filename)
            
            self.status_var.set("Commandes sauvegardées avec succès")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la sauvegarde des commandes:\n{str(e)}")

    def save_consumption(self):
        """Enregistre une nouvelle consommation dans le fichier Excel"""
        product_type = self.product_type.get()
        ref = self.combo_ref.get()
        
        if not ref:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une référence de produit")
            return
            
        # Validation de la consommation
        try:
            consommation = float(self.entry_consommation.get())
            if consommation <= 0:
                raise ValueError
        except:
            messagebox.showwarning("Erreur", "Veuillez entrer une valeur numérique valide (> 0)")
            return
            
        # Vérification du stock suffisant
        if product_type == "Colorant":
            current_stock = self.stocks.get(ref, 0.0)
        else:
            # Pour les produits auxiliaires, récupérer le stock depuis Excel
            current_stock = 0.0
            try:
                sheet = self.wb["Produits auxiliaires"]
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == ref:
                        current_stock = sheet.cell(row=row, column=6).value or 0.0
                        try:
                            current_stock = float(current_stock)
                        except:
                            current_stock = 0.0
                        break
            except:
                pass
        
        if consommation > current_stock:
            messagebox.showwarning("Erreur", 
                               f"Stock insuffisant! Stock actuel: {current_stock:.2f} kg\n"
                               f"Consommation demandée: {consommation:.2f} kg")
            return
            
        # Validation de la date
        date_str = self.entry_date.get()
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showwarning("Erreur", "Format de date invalide. Utilisez AAAA-MM-JJ")
            return
        
        try:
            # Ajout dans la feuille Consommation
            new_row = self.sheet_consommation.max_row + 1
            self.sheet_consommation.cell(row=new_row, column=1, value=ref)
            self.sheet_consommation.cell(row=new_row, column=2, value=date_str)
            self.sheet_consommation.cell(row=new_row, column=3, value=consommation)
            
            # Mise à jour du stock réel
            nouveau_stock = current_stock - consommation
            
            if product_type == "Colorant":
                self.stocks[ref] = nouveau_stock
                
                # Trouver la ligne dans Liste des articles2
                for row in range(2, self.sheet_articles.max_row + 1):
                    cell_ref = self.sheet_articles.cell(row=row, column=1).value
                    if cell_ref and str(cell_ref) == ref:
                        # Mettre à jour le stock réel (colonne F/6)
                        self.sheet_articles.cell(row=row, column=6).value = nouveau_stock
                        
                        # Mise à jour de l'alerte de stock (colonne H/8)
                        stock_min = self.stock_min.get(ref, 0.0)
                        alerte = "vrai" if nouveau_stock < stock_min else "faux"
                        self.sheet_articles.cell(row=row, column=8).value = alerte
                        break
            else:
                # Mise à jour du produit auxiliaire
                sheet = self.wb["Produits auxiliaires"]
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == ref:
                        # Mettre à jour le stock réel (colonne F/6)
                        sheet.cell(row=row, column=6).value = nouveau_stock
                        
                        # Mise à jour de l'alerte de stock (colonne H/8)
                        try:
                            stock_min = float(sheet.cell(row=row, column=4).value)
                        except:
                            stock_min = 0.0
                        alerte = "vrai" if nouveau_stock < stock_min else "faux"
                        sheet.cell(row=row, column=8).value = alerte
                        break
            
            # Sauvegarde du fichier
            self.wb.save(self.filename)
            
            # Ajouter à l'historique
            self.consumption_history.append({
                'ref': ref,
                'date': date_str,
                'qty': consommation,
                'id': new_row
            })
            
            # Mise à jour de l'interface
            self.label_stock.config(text=f"{nouveau_stock:.2f} kg")
            self.entry_consommation.delete(0, tk.END)
            self.update_history_tree()
            self.check_stock_alerts()
            self.update_report_table()
            self.update_indicators()
            
            # Afficher une alerte si le stock passe sous le minimum
            if nouveau_stock < stock_min:
                messagebox.showwarning("Alerte Stock", 
                                   f"Attention! Le stock de {ref} est tombé à {nouveau_stock:.2f} kg, "
                                   f"ce qui est en dessous du stock minimal de {stock_min:.2f} kg.")
            
            messagebox.showinfo("Succès", "Consommation enregistrée avec succès!")
            self.status_var.set(f"Consommation de {consommation:.2f} kg enregistrée pour {ref}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement:\n{str(e)}")

    def update_initial_stock(self):
        """Met à jour le stock initial dans Excel"""
        product_type = self.stock_product_type.get()
        ref = self.combo_stock_ref.get()
        
        if not ref:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une référence de produit")
            return
            
        # Validation du nouveau stock
        try:
            new_stock = float(self.entry_new_stock.get())
            if new_stock < 0:
                raise ValueError
        except:
            messagebox.showwarning("Erreur", "Veuillez entrer une valeur numérique valide (≥ 0)")
            return
            
        try:
            if product_type == "Colorant":
                # Trouver la ligne dans Liste des articles2
                for row in range(2, self.sheet_articles.max_row + 1):
                    cell_ref = self.sheet_articles.cell(row=row, column=1).value
                    if cell_ref and str(cell_ref) == ref:
                        # Mettre à jour le stock initial (colonne C/3)
                        self.sheet_articles.cell(row=row, column=3).value = new_stock
                        
                        # Recalculer le stock réel (F = C - E)
                        consommation = self.sheet_articles.cell(row=row, column=5).value or 0
                        try:
                            consommation = float(consommation)
                        except (TypeError, ValueError):
                            consommation = 0.0
                        
                        nouveau_stock_reel = new_stock - consommation
                        self.sheet_articles.cell(row=row, column=6).value = nouveau_stock_reel
                        
                        # Mise à jour des données internes
                        self.stock_initial[ref] = new_stock
                        self.stocks[ref] = nouveau_stock_reel
                        
                        # Mise à jour de l'alerte de stock
                        stock_min = self.stock_min.get(ref, 0.0)
                        alerte = "vrai" if nouveau_stock_reel < stock_min else "faux"
                        self.sheet_articles.cell(row=row, column=8).value = alerte
                        break
            else:
                # Mise à jour du produit auxiliaire
                sheet = self.wb["Produits auxiliaires"]
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == ref:
                        # Mettre à jour le stock initial (colonne C/3)
                        sheet.cell(row=row, column=3).value = new_stock
                        
                        # Recalculer le stock réel (F = C - E)
                        consommation = sheet.cell(row=row, column=5).value or 0
                        try:
                            consommation = float(consommation)
                        except (TypeError, ValueError):
                            consommation = 0.0
                        
                        nouveau_stock_reel = new_stock - consommation
                        sheet.cell(row=row, column=6).value = nouveau_stock_reel
                        
                        # Mise à jour de l'alerte de stock
                        try:
                            stock_min = float(sheet.cell(row=row, column=4).value)
                        except:
                            stock_min = 0.0
                        alerte = "vrai" if nouveau_stock_reel < stock_min else "faux"
                        sheet.cell(row=row, column=8).value = alerte
                        break
            
            # Sauvegarde du fichier
            self.wb.save(self.filename)
            
            # Mise à jour de l'interface
            self.label_current_stock.config(text=f"{new_stock:.2f} kg")
            self.update_stock_display()
            self.check_stock_alerts()
            self.update_report_table()
            self.update_indicators()
            
            messagebox.showinfo("Succès", "Stock initial mis à jour avec succès!")
            self.status_var.set(f"Stock initial de {ref} mis à jour: {new_stock:.2f} kg")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la mise à jour du stock:\n{str(e)}")

    def add_new_product(self):
        """Ajoute un nouveau produit au fichier Excel"""
        product_type = self.new_product_type.get()
        ref = self.entry_new_ref.get().strip()
        name = self.entry_new_name.get().strip()
        init_stock = self.entry_new_init_stock.get().strip()
        min_stock = self.entry_new_min_stock.get().strip()
        
        # Validation des champs
        if not ref or not name:
            messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
            return
            
        # Validation des valeurs numériques
        try:
            init_stock = float(init_stock) if init_stock else 0.0
            min_stock = float(min_stock) if min_stock else 0.0
        except ValueError:
            messagebox.showwarning("Erreur", "Les valeurs de stock doivent être numériques")
            return
            
        try:
            if product_type == "Colorant":
                if ref in self.colorants:
                    messagebox.showwarning("Erreur", "Cette référence existe déjà")
                    return
                    
                # Trouver la première ligne vide
                new_row = self.sheet_articles.max_row + 1
                
                # Ajouter le nouveau colorant
                self.sheet_articles.cell(row=new_row, column=1, value=ref)
                self.sheet_articles.cell(row=new_row, column=2, value=name)
                self.sheet_articles.cell(row=new_row, column=3, value=init_stock)
                self.sheet_articles.cell(row=new_row, column=4, value=min_stock)
                
                # Ajouter les formules
                # CONSOMMATION =SUMIF(Consommation!A2:A{self.sheet_consommation.max_row},A{new_row},Consommation!C2:C500)
                self.sheet_articles.cell(row=new_row, column=5, 
                                        value=f'=SUMIF(Consommation!A2:A{self.sheet_consommation.max_row},A{new_row},Consommation!C2:C500)')
                
                # STOCK REEL =C3-E3
                self.sheet_articles.cell(row=new_row, column=6, value=f'=C{new_row}-E{new_row}')
                
                # ALERTE DE STOCK =IF(D3>=F3,"faux","vrai")
                self.sheet_articles.cell(row=new_row, column=8, value=f'=IF(D{new_row}>=F{new_row},"faux","vrai")')
                
            else:
                # Produit auxiliaire
                sheet = self.wb["Produits auxiliaires"]
                # Vérifier si la référence existe déjà
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=row, column=1).value) == ref:
                        messagebox.showwarning("Erreur", "Cette référence existe déjà")
                        return
                
                new_row = sheet.max_row + 1
                sheet.cell(row=new_row, column=1, value=ref)
                sheet.cell(row=new_row, column=2, value=name)
                sheet.cell(row=new_row, column=3, value=init_stock)
                sheet.cell(row=new_row, column=4, value=min_stock)
                
                # Ajouter les formules
                # CONSOMMATION =SUMIF(Consommation!A2:A{self.sheet_consommation.max_row},A{new_row},Consommation!C2:C500)
                sheet.cell(row=new_row, column=5, 
                          value=f'=SUMIF(Consommation!A2:A{self.sheet_consommation.max_row},A{new_row},Consommation!C2:C500)')
                
                # STOCK REEL =C3-E3
                sheet.cell(row=new_row, column=6, value=f'=C{new_row}-E{new_row}')
                
                # ALERTE DE STOCK =IF(D3>=F3,"faux","vrai")
                sheet.cell(row=new_row, column=8, value=f'=IF(D{new_row}>=F{new_row},"faux","vrai")')
            
            # Sauvegarde du fichier
            self.wb.save(self.filename)
            
            # Recharger les données
            self.load_data()
            
            # Mettre à jour les combobox
            self.update_product_list()
            self.update_stock_product_list()
            
            # Sélectionner le nouveau produit
            self.combo_ref.set(ref)
            self.combo_stock_ref.set(ref)
            self.update_stock_display()
            self.update_stock_info()
            self.check_stock_alerts()
            self.update_report_table()
            self.update_indicators()
            
            # Vider les champs
            self.entry_new_ref.delete(0, tk.END)
            self.entry_new_name.delete(0, tk.END)
            self.entry_new_init_stock.delete(0, tk.END)
            self.entry_new_min_stock.delete(0, tk.END)
            
            messagebox.showinfo("Succès", "Nouveau produit ajouté avec succès!")
            self.status_var.set(f"Produit ajouté: {ref} - {name}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout du produit:\n{str(e)}")

    def create_auxiliary_tab(self):
        """Crée l'onglet pour les produits auxiliaires"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🧪 Produits Auxiliaires")
        
        # Frame principale avec padding
        main_frame = ttk.Frame(tab, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Titre
        header = ttk.Label(main_frame, text="Gestion des Produits Auxiliaires", 
                          font=("Segoe UI", 12, "bold"), anchor="center")
        header.pack(fill=tk.X, pady=(0, 15))
        
        # Frame avec deux colonnes
        dual_frame = ttk.Frame(main_frame)
        dual_frame.pack(fill=tk.BOTH, expand=True)
        
        # Colonne gauche - Ajout de produit
        left_frame = ttk.LabelFrame(dual_frame, text="➕ Ajouter un Nouveau Produit")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10), pady=5)
        
        # Formulaire
        form_frame = ttk.Frame(left_frame, padding=10)
        form_frame.pack(fill=tk.X)
        
        # Champs pour nouveau produit
        ttk.Label(form_frame, text="ID Produit:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.entry_aux_id = ttk.Entry(form_frame, width=30)
        self.entry_aux_id.grid(row=0, column=1, sticky="we", padx=5, pady=5)
        
        ttk.Label(form_frame, text="Nom du Produit:", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.entry_aux_name = ttk.Entry(form_frame, width=30)
        self.entry_aux_name.grid(row=1, column=1, sticky="we", padx=5, pady=5)
        
        ttk.Label(form_frame, text="Stock Initial:", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.entry_aux_stock = ttk.Entry(form_frame, width=15)
        self.entry_aux_stock.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        self.entry_aux_stock.insert(0, "0.0")
        
        ttk.Label(form_frame, text="Stock Minimal:", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.entry_aux_min = ttk.Entry(form_frame, width=15)
        self.entry_aux_min.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        self.entry_aux_min.insert(0, "0.0")
        
        # Bouton d'ajout
        btn_frame = ttk.Frame(left_frame, padding=10)
        btn_frame.pack(fill=tk.X)
        self.btn_add_aux = ttk.Button(btn_frame, text="➕ Ajouter Produit", 
                                    command=self.add_auxiliary_product, style="Accent.TButton")
        self.btn_add_aux.pack(pady=5)
        
        # Colonne droite - Liste des produits
        right_frame = ttk.LabelFrame(dual_frame, text="📋 Liste des Produits Auxiliaires")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=5)
        
        # Tableau des produits
        columns = ("id", "name", "stock", "min_stock")
        self.aux_tree = ttk.Treeview(right_frame, columns=columns, show="headings")
        
        # Configuration des colonnes
        self.aux_tree.heading("id", text="ID Produit")
        self.aux_tree.heading("name", text="Nom")
        self.aux_tree.heading("stock", text="Stock")
        self.aux_tree.heading("min_stock", text="Stock Min")
        
        self.aux_tree.column("id", width=100, anchor="center")
        self.aux_tree.column("name", width=200, anchor="w")
        self.aux_tree.column("stock", width=100, anchor="e")
        self.aux_tree.column("min_stock", width=100, anchor="e")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.aux_tree.yview)
        self.aux_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement des éléments
        self.aux_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Boutons d'action
        btn_frame = ttk.Frame(right_frame, padding=5)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(btn_frame, text="🔄 Actualiser", command=self.load_auxiliary_data).pack(side=tk.LEFT, padx=2)

    def create_indicators_tab(self):
        """Crée l'onglet pour les indicateurs"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📈 Indicateurs")
        
        # Frame principale avec padding
        main_frame = ttk.Frame(tab, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Titre
        header = ttk.Label(main_frame, text="Indicateurs de Performance et Consommation", 
                          font=("Segoe UI", 12, "bold"), anchor="center")
        header.pack(fill=tk.X, pady=(0, 15))
        
        # Frame avec deux colonnes
        dual_frame = ttk.Frame(main_frame)
        dual_frame.pack(fill=tk.BOTH, expand=True)
        
        # Colonne gauche - Statistiques
        left_frame = ttk.LabelFrame(dual_frame, text="📊 Statistiques Clés")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10), pady=5)
        
        # KPI Cards
        kpi_frame = ttk.Frame(left_frame, padding=10)
        kpi_frame.pack(fill=tk.BOTH, expand=True)
        
        # Card 1: Total des commandes
        card1 = ttk.Frame(kpi_frame, relief="solid", borderwidth=1, padding=10)
        card1.pack(fill=tk.X, pady=5)
        
        ttk.Label(card1, text="Commandes Totales", font=("Segoe UI", 9)).pack(anchor="w")
        self.total_cmd_kpi = ttk.Label(card1, text="0", font=("Segoe UI", 24, "bold"), foreground="#3b5998")
        self.total_cmd_kpi.pack(anchor="center", pady=5)
        
        # Card 2: Commandes traitées
        card2 = ttk.Frame(kpi_frame, relief="solid", borderwidth=1, padding=10)
        card2.pack(fill=tk.X, pady=5)
        
        ttk.Label(card2, text="Commandes Traitées", font=("Segoe UI", 9)).pack(anchor="w")
        self.traitees_kpi = ttk.Label(card2, text="0", font=("Segoe UI", 24, "bold"), foreground="#4CAF50")
        self.traitees_kpi.pack(anchor="center", pady=5)
        
        # Card 3: Taux de traitement
        card3 = ttk.Frame(kpi_frame, relief="solid", borderwidth=1, padding=10)
        card3.pack(fill=tk.X, pady=5)
        
        ttk.Label(card3, text="Taux de Traitement", font=("Segoe UI", 9)).pack(anchor="w")
        self.taux_kpi = ttk.Label(card3, text="0%", font=("Segoe UI", 24, "bold"), foreground="#FF9800")
        self.taux_kpi.pack(anchor="center", pady=5)
        
        # Bouton d'actualisation
        btn_frame = ttk.Frame(left_frame, padding=10)
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="🔄 Actualiser les Indicateurs", 
                  command=self.update_indicators, style="Accent.TButton").pack()
        
        # Colonne droite - Consommation
        right_frame = ttk.LabelFrame(dual_frame, text="🔥 Top 10 des Produits les Plus Consommés")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=5)
        
        # Tableau des consommations
        columns = ("rank", "name", "total_cons", "priority", "type")
        self.top_cons_tree = ttk.Treeview(right_frame, columns=columns, show="headings")
        
        # Configuration des colonnes
        self.top_cons_tree.heading("rank", text="#")
        self.top_cons_tree.heading("name", text="Produit")
        self.top_cons_tree.heading("total_cons", text="Consommation Totale")
        self.top_cons_tree.heading("priority", text="Priorité")
        self.top_cons_tree.heading("type", text="Type")
        
        self.top_cons_tree.column("rank", width=50, anchor="center")
        self.top_cons_tree.column("name", width=200, anchor="w")
        self.top_cons_tree.column("total_cons", width=150, anchor="e")
        self.top_cons_tree.column("priority", width=100, anchor="center")
        self.top_cons_tree.column("type", width=120, anchor="center")
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=self.top_cons_tree.yview)
        self.top_cons_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement des éléments
        self.top_cons_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bouton d'export
        btn_frame = ttk.Frame(right_frame, padding=5)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

    def update_indicators(self):
        """Met à jour les indicateurs de performance"""
        # Calculer la consommation totale par produit
        consumption_totals = {}
        for item in self.consumption_history:
            ref = item['ref']
            consumption_totals[ref] = consumption_totals.get(ref, 0) + item['qty']
        
        # Ajouter les noms et types de produits
        products_with_info = []
        for ref, total in consumption_totals.items():
            # Déterminer le type de produit
            product_type = "Colorant" if ref in self.colorants else "Produit auxiliaire"
            
            # Trouver le nom du produit
            if product_type == "Colorant":
                name = self.colorant_names.get(ref, ref)
            else:
                name = ref  # Par défaut
                try:
                    sheet = self.wb["Produits auxiliaires"]
                    for row in range(2, sheet.max_row + 1):
                        if str(sheet.cell(row=row, column=1).value) == ref:
                            name = sheet.cell(row=row, column=2).value or ref
                            break
                except:
                    pass
            
            products_with_info.append({
                'ref': ref,
                'name': name,
                'total': total,
                'type': product_type
            })
        
        # Trier par consommation décroissante
        sorted_consumption = sorted(products_with_info, key=lambda x: x['total'], reverse=True)
        
        # Mettre à jour le top 10
        for item in self.top_cons_tree.get_children():
            self.top_cons_tree.delete(item)
            
        for rank, product in enumerate(sorted_consumption[:10], 1):
            priority = "Élevée" if rank <= 3 else "Moyenne" if rank <= 7 else "Basse"
            self.top_cons_tree.insert("", "end", values=(
                rank, 
                product['name'], 
                f"{product['total']:.2f} kg", 
                priority,
                product['type']
            ))
            
            # Colorer les priorités élevées
            if rank <= 3:
                self.top_cons_tree.item(self.top_cons_tree.get_children()[-1], tags=("high",))
        
        # Configurer le style pour les priorités
        self.top_cons_tree.tag_configure("high", background="#fff9c4")
        
        # Mettre à jour les KPI
        self.total_cmd_kpi.config(text=str(self.total_commandes))
        self.traitees_kpi.config(text=str(self.commandes_traitees))
        self.taux_kpi.config(text=f"{self.taux_commandes:.1f}%")
        
        self.status_var.set("Indicateurs mis à jour")

    def edit_consumption(self):
        """Modifie une consommation sélectionnée"""
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une consommation à modifier")
            return
            
        item = self.history_tree.item(selected[0])
        values = item['values']
        row_id = values[5]  # ID de la ligne dans Excel
        
        # Trouver la consommation
        consumption = next((c for c in self.consumption_history if c['id'] == row_id), None)
        if not consumption:
            messagebox.showwarning("Erreur", "Consommation introuvable")
            return
            
        dialog = tk.Toplevel(self)
        dialog.title("Modifier Consommation")
        dialog.geometry("400x300")
        dialog.transient(self)
        dialog.grab_set()
        
        # Formulaire
        form_frame = ttk.Frame(dialog, padding=15)
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Référence (non modifiable)
        ttk.Label(form_frame, text="Référence:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        ref_label = ttk.Label(form_frame, text=consumption['ref'], font=("Segoe UI", 10, "bold"))
        ref_label.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        # Type (non modifiable)
        ttk.Label(form_frame, text="Type:", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        type_label = ttk.Label(form_frame, text="Colorant" if consumption['ref'] in self.colorants else "Produit auxiliaire", 
                             font=("Segoe UI", 10))
        type_label.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        # Date
        ttk.Label(form_frame, text="Date:", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        date_entry = ttk.Entry(form_frame, width=15)
        date_entry.insert(0, consumption['date'])
        date_entry.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        # Quantité
        ttk.Label(form_frame, text="Quantité (kg):", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        qty_entry = ttk.Entry(form_frame, width=15)
        qty_entry.insert(0, str(consumption['qty']))
        qty_entry.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Boutons
        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(fill=tk.X)
        
        def save_changes():
            # Validation des données
            try:
                new_date = date_entry.get()
                datetime.strptime(new_date, '%Y-%m-%d')
                new_qty = float(qty_entry.get())
                if new_qty <= 0:
                    raise ValueError
            except:
                messagebox.showwarning("Erreur", "Veuillez entrer des données valides")
                return
                
            # Mettre à jour la consommation
            consumption['date'] = new_date
            consumption['qty'] = new_qty
            
            # Mettre à jour Excel
            try:
                self.sheet_consommation.cell(row=row_id, column=2, value=new_date)
                self.sheet_consommation.cell(row=row_id, column=3, value=new_qty)
                self.wb.save(self.filename)
                
                # Recalculer les stocks
                self.load_data()
                self.update_stock_display()
                self.update_history_tree()
                self.update_report_table()
                self.update_indicators()
                self.check_stock_alerts()
                
                messagebox.showinfo("Succès", "Consommation modifiée avec succès")
                self.status_var.set(f"Consommation du {new_date} pour {consumption['ref']} modifiée")
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la modification:\n{str(e)}")
        
        ttk.Button(btn_frame, text="Enregistrer", command=save_changes, style="Accent.TButton").pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Annuler", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def delete_consumption(self):
        """Supprime une consommation sélectionnée"""
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une consommation à supprimer")
            return
            
        item = self.history_tree.item(selected[0])
        values = item['values']
        row_id = values[5]  # ID de la ligne dans Excel
        
        # Confirmation
        if not messagebox.askyesno("Confirmation", 
                                  f"Voulez-vous vraiment supprimer cette consommation?\n"
                                  f"Référence: {values[1]}\n"
                                  f"Date: {values[0]}\n"
                                  f"Quantité: {values[3]} kg"):
            return
            
        try:
            # Supprimer la ligne dans Excel
            self.sheet_consommation.delete_rows(row_id)
            self.wb.save(self.filename)
            
            # Mettre à jour les données
            self.consumption_history = [c for c in self.consumption_history if c['id'] != row_id]
            
            # Recalculer les stocks
            self.load_data()
            self.update_stock_display()
            self.update_history_tree()
            self.update_report_table()
            self.update_indicators()
            self.check_stock_alerts()
            
            messagebox.showinfo("Succès", "Consommation supprimée avec succès")
            self.status_var.set(f"Consommation du {values[0]} pour {values[1]} supprimée")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la suppression:\n{str(e)}")

    def add_auxiliary_product(self):
        """Ajoute un nouveau produit auxiliaire"""
        product_id = self.entry_aux_id.get().strip()
        name = self.entry_aux_name.get().strip()
        stock = self.entry_aux_stock.get().strip()
        min_stock = self.entry_aux_min.get().strip()
        
        # Validation des données
        if not product_id or not name:
            messagebox.showwarning("Erreur", "L'ID et le nom sont obligatoires")
            return
            
        try:
            stock = float(stock) if stock else 0.0
            min_stock = float(min_stock) if min_stock else 0.0
        except ValueError:
            messagebox.showwarning("Erreur", "Les valeurs de stock doivent être numériques")
            return
            
        try:
            sheet = self.wb["Produits auxiliaires"]
            
            # Vérifier si l'ID existe déjà
            for row in range(2, sheet.max_row + 1):
                if str(sheet.cell(row=row, column=1).value) == product_id:
                    messagebox.showwarning("Erreur", "Cet ID de produit existe déjà")
                    return
            
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=1, value=product_id)
            sheet.cell(row=new_row, column=2, value=name)
            sheet.cell(row=new_row, column=3, value=stock)
            sheet.cell(row=new_row, column=4, value=min_stock)
            sheet.cell(row=new_row, column=6, value=f"=C{new_row}-E{new_row}")
            sheet.cell(row=new_row, column=8, value=f'=IF(D{new_row}>=F{new_row},"faux","vrai")')
            
            self.wb.save(self.filename)
            
            # Vider les champs
            self.entry_aux_id.delete(0, tk.END)
            self.entry_aux_name.delete(0, tk.END)
            self.entry_aux_stock.delete(0, tk.END)
            self.entry_aux_min.delete(0, tk.END)
            
            messagebox.showinfo("Succès", "Produit auxiliaire ajouté avec succès")
            self.load_auxiliary_data()
            self.update_product_list()
            self.update_stock_product_list()
            self.update_report_table()
            self.update_indicators()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout:\n{str(e)}")

    def load_auxiliary_data(self):
        """Charge les produits auxiliaires"""
        try:
            # Effacer l'arbre
            for item in self.aux_tree.get_children():
                self.aux_tree.delete(item)
                
            sheet = self.wb["Produits auxiliaires"]
            for row in range(2, sheet.max_row + 1):
                product_id = sheet.cell(row=row, column=1).value
                if product_id:
                    name = sheet.cell(row=row, column=2).value or ""
                    
                    try:
                        stock = float(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3).value is not None else 0.0
                    except:
                        stock = 0.0
                    
                    try:
                        min_stock = float(sheet.cell(row=row, column=4).value) if sheet.cell(row=row, column=4).value is not None else 0.0
                    except:
                        min_stock = 0.0
                    
                    self.aux_tree.insert("", "end", values=(
                        product_id, name, f"{stock:.2f}", f"{min_stock:.2f}"
                    ))
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur de chargement:\n{str(e)}")

    def export_to_excel(self):
        """Exporte le rapport actuel vers un nouveau fichier Excel"""
        try:
            # Demander le nom du fichier
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
                title="Enregistrer le rapport"
            )
            
            if not filepath:
                return
                
            # Créer un nouveau classeur
            export_wb = openpyxl.Workbook()
            sheet = export_wb.active
            sheet.title = "Rapport de Stock"
            
            # En-têtes
            headers = ["Référence", "Nom", "Stock Initial", "Stock Réel", "Stock Minimal", "Statut", "Type"]
            sheet.append(headers)
            
            # Données
            for item in self.report_tree.get_children():
                values = self.report_tree.item(item)['values']
                sheet.append(values)
            
            # Sauvegarder
            export_wb.save(filepath)
            
            messagebox.showinfo("Export Réussi", 
                               f"Le rapport a été exporté avec succès dans:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Erreur d'Export", f"Erreur lors de l'exportation:\n{str(e)}")


if __name__ == "__main__":
    app = StockApp()
    app.mainloop()